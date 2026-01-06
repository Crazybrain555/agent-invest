"""
Excel 报告生成

按指定列顺序输出，内存全 float → 写 Excel 再做展示格式
"""

import logging
from pathlib import Path
from typing import Union

import pandas as pd

from backtest.backtest_result_pipeline.types import AggregatedTables
from backtest.backtest_result_pipeline.io.atomic_write import atomic_write_df

logger = logging.getLogger(__name__)


# 列定义
SUMMARY_COLUMNS = [
    "股票池", "时期", "策略", "总收益率", "基准收益率", "超额收益率",
    "夏普比率", "最大回撤", "Calmar比率", "胜率", "IC均值", "IC胜率"
]

# 导出列顺序（总体表现 & 年度表现）
# 说明：
# - 口径：超额收益率使用“算术差”(strategy - benchmark)
# - 年度表现：连续 NAV 切片（含上一交易日基准）
# - 展示顺序：把基准/超额放前面，便于快速对比
OVERALL_EXPORT_COLUMNS = [
    "pool_code",
    "strategy_name",
    "total_return",
    "benchmark_code",
    "benchmark_return",
    "excess_return",
    "annual_return",
    "volatility",
    "sharpe_ratio",
    "max_drawdown",
    "calmar_ratio",
    "hit_rate",
    "profit_loss_ratio",
    "var_95",
    "cvar_95",
    "mean_ic",
    "ic_std",
    "ic_hit_rate",
    "factor_return_total",
    "factor_return_mean",
    "factor_return_t_stat",
    "turnover_mean",
    "turnover_total",
]

OVERALL_EXPORT_COLUMNS_CN = [
    "股票池",
    "策略名称",
    "总收益率",
    "基准代码",
    "基准收益率",
    "超额收益率",
    "年化收益率",
    "波动率",
    "夏普比率",
    "最大回撤",
    "Calmar比率",
    "胜率",
    "盈亏比",
    "VaR(95%)",
    "CVaR(95%)",
    "IC均值",
    "IC标准差",
    "IC胜率",
    "因子收益率总和",
    "因子收益率均值",
    "因子收益率T值",
    "平均换手率",
    "总换手率",
]

OVERALL_RENAME = dict(zip(OVERALL_EXPORT_COLUMNS, OVERALL_EXPORT_COLUMNS_CN))

YEARLY_EXPORT_COLUMNS = ["pool_code", "year", "strategy_name"] + [
    c for c in OVERALL_EXPORT_COLUMNS if c not in {"strategy_name", "pool_code"}
]
YEARLY_RENAME = {"year": "年份", "strategy_name": "策略名称", "pool_code": "股票池"} | OVERALL_RENAME


def _format_percentage(x):
    """格式化为百分比"""
    if pd.isna(x):
        return ""
    return f"{x:.2%}"


def _format_float(x, decimals=4):
    """格式化浮点数"""
    if pd.isna(x):
        return ""
    return f"{x:.{decimals}f}"


def _apply_excel_formats(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame):
    """应用 Excel 格式"""
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    
    # 定义格式
    pct_format = workbook.add_format({"num_format": "0.00%"})
    float_format = workbook.add_format({"num_format": "0.0000"})
    float2_format = workbook.add_format({"num_format": "0.00"})
    
    # 百分比列
    pct_cols = [
        "总收益率", "基准收益率", "超额收益率", "年化收益率", "波动率", "最大回撤", "胜率",
        "IC胜率", "VaR(95%)", "CVaR(95%)", "平均换手率", "总换手率",
        "total_return", "benchmark_return", "excess_return", "annual_return", "volatility", "max_drawdown",
        "hit_rate", "ic_hit_rate", "var_95", "cvar_95", "turnover_mean", "turnover_total",
    ]
    
    # 两位小数列
    float2_cols = ["夏普比率", "Calmar比率", "盈亏比", "sharpe_ratio", "calmar_ratio", "profit_loss_ratio"]
    
    for col_idx, col_name in enumerate(df.columns):
        if col_name in pct_cols:
            worksheet.set_column(col_idx, col_idx, 12, pct_format)
        elif col_name in float2_cols:
            worksheet.set_column(col_idx, col_idx, 12, float2_format)
        else:
            worksheet.set_column(col_idx, col_idx, 14, float_format)


def _apply_openpyxl_formats(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame) -> None:
    """openpyxl 引擎下的格式化（避免依赖 xlsxwriter）"""
    try:
        from openpyxl.utils import get_column_letter
    except Exception:
        # openpyxl 不可用时直接跳过
        return

    worksheet = writer.sheets[sheet_name]

    pct_cols = {
        "总收益率", "基准收益率", "超额收益率", "年化收益率", "波动率", "最大回撤", "胜率",
        "IC胜率", "VaR(95%)", "CVaR(95%)", "平均换手率", "总换手率",
    }
    float2_cols = {"夏普比率", "Calmar比率", "盈亏比"}

    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)

        if col_name in pct_cols:
            number_format = "0.00%"
            width = 12
        elif col_name in float2_cols:
            number_format = "0.00"
            width = 12
        elif col_name in {"年份"}:
            number_format = "0"
            width = 8
        else:
            number_format = "0.0000"
            width = 14

        worksheet.column_dimensions[col_letter].width = width

        # 从第 2 行开始应用格式（第 1 行是表头）
        for row_idx in range(2, len(df) + 2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            # 仅对数值型或空值应用 number_format；字符串保持默认
            if cell.value is None:
                continue
            if isinstance(cell.value, (int, float)):
                cell.number_format = number_format


def write_excel_report(
    tables: AggregatedTables,
    output_path: Union[str, Path],
    no_overwrite: bool = True
) -> Path:
    """
    写入 Excel 报告
    
    Args:
        tables: 聚合后的表格
        output_path: 输出路径
        no_overwrite: 是否禁止覆盖
    
    Returns:
        实际写入的路径
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    logger.info(f"生成 Excel 报告: {output_path}")
    
    # 准备数据
    summary_df = tables.summary.copy()
    overall_df_raw = tables.overall.copy()
    yearly_df_raw = tables.yearly.copy()

    # ========== 总体表现：重排 + 中文列名 ==========
    overall_existing_cols = [c for c in OVERALL_EXPORT_COLUMNS if c in overall_df_raw.columns]
    overall_df_display = overall_df_raw.reindex(columns=overall_existing_cols).rename(columns=OVERALL_RENAME)

    # ========== 年度表现：补齐基准代码 + 重排 + 中文列名 ==========
    yearly_df_work = yearly_df_raw.copy()

    # 补齐 benchmark_code（年度表中重复填充，便于对齐 overall/计划模板）
    if "benchmark_code" not in yearly_df_work.columns and {"pool_code", "strategy_name", "benchmark_code"} <= set(overall_df_raw.columns):
        code_lookup = overall_df_raw[["pool_code", "strategy_name", "benchmark_code"]].drop_duplicates()
        if {"pool_code", "strategy_name"} <= set(yearly_df_work.columns):
            yearly_df_work = yearly_df_work.merge(
                code_lookup,
                on=["pool_code", "strategy_name"],
                how="left"
            )

    yearly_existing_cols = [c for c in YEARLY_EXPORT_COLUMNS if c in yearly_df_work.columns]
    yearly_df_display = yearly_df_work.reindex(columns=yearly_existing_cols).rename(columns=YEARLY_RENAME)
    
    # 写入 Excel（使用 xlsxwriter 引擎以支持格式化）
    try:
        with pd.ExcelWriter(output_path, engine="xlsxwriter") as writer:
            # Sheet 1: 核心指标汇总
            summary_df.to_excel(writer, sheet_name="核心指标汇总", index=False)
            _apply_excel_formats(writer, "核心指标汇总", summary_df)
            
            # Sheet 2: 总体表现
            overall_df_display.to_excel(writer, sheet_name="总体表现", index=False)
            _apply_excel_formats(writer, "总体表现", overall_df_display)
            
            # Sheet 3: 年度表现
            if not yearly_df_display.empty:
                yearly_df_display.to_excel(writer, sheet_name="年度表现", index=False)
                _apply_excel_formats(writer, "年度表现", yearly_df_display)
        
        logger.info(f"   Excel 报告已生成: {output_path}")
        return output_path
    
    except ImportError:
        # 如果没有 xlsxwriter，使用 openpyxl
        logger.warning("xlsxwriter 未安装，使用 openpyxl（格式化可能不完整）")
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="核心指标汇总", index=False)
            _apply_openpyxl_formats(writer, "核心指标汇总", summary_df)

            overall_df_display.to_excel(writer, sheet_name="总体表现", index=False)
            _apply_openpyxl_formats(writer, "总体表现", overall_df_display)
            if not yearly_df_display.empty:
                yearly_df_display.to_excel(writer, sheet_name="年度表现", index=False)
                _apply_openpyxl_formats(writer, "年度表现", yearly_df_display)
        
        return output_path
