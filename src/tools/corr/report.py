from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Optional, Tuple

import pandas as pd
import yaml


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def write_summary(path: Path, summary: Dict[str, Any]) -> None:
    ensure_dir(path.parent)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(summary, handle, indent=2, ensure_ascii=True)


def write_parquet(path: Path, df: pd.DataFrame) -> None:
    ensure_dir(path.parent)
    df.to_parquet(path, index=False)


def write_recommendation(path: Path, data: Dict[str, Any]) -> None:
    ensure_dir(path.parent)
    with path.open("w", encoding="utf-8") as handle:
        yaml.safe_dump(data, handle, allow_unicode=False, sort_keys=False)


def write_text(path: Path, text: str) -> None:
    ensure_dir(path.parent)
    with path.open("w", encoding="utf-8") as handle:
        handle.write(text)


def write_excel(
    path: Path,
    df: pd.DataFrame,
    sheet_name: str,
    index: bool = False,
    freeze_panes: Optional[str] = None,
    color_scale_range: Optional[Tuple[int, int, int, int]] = None,
) -> None:
    ensure_dir(path.parent)
    try:
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise RuntimeError("openpyxl is required for Excel output") from exc

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=index)
        workbook = writer.book
        worksheet = workbook[sheet_name]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="366092")
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        if freeze_panes:
            worksheet.freeze_panes = freeze_panes

        if color_scale_range:
            start_row, start_col, end_row, end_col = color_scale_range
            if end_row >= start_row and end_col >= start_col:
                start_col_letter = get_column_letter(start_col)
                end_col_letter = get_column_letter(end_col)
                cell_range = f"{start_col_letter}{start_row}:{end_col_letter}{end_row}"
                rule = ColorScaleRule(
                    start_type="num",
                    start_value=-1,
                    start_color="4575B4",
                    mid_type="num",
                    mid_value=0,
                    mid_color="FFFFBF",
                    end_type="num",
                    end_value=1,
                    end_color="D73027",
                )
                worksheet.conditional_formatting.add(cell_range, rule)

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                value = cell.value
                if value is None:
                    continue
                max_length = max(max_length, len(str(value)))
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
